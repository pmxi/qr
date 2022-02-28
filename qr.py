from openpyxl import load_workbook
import time
import cv2
import numpy as np
from pyzbar.pyzbar import decode


def decoder(image):

    gray_img = cv2.cvtColor(image, 0)
    barcode = decode(gray_img)

    for obj in barcode:
        points = obj.polygon
        (x, y, w, h) = obj.rect
        pts = np.array(points, np.int32)
        pts = pts.reshape((-1, 1, 2))
        cv2.polylines(image, [pts], True, (0, 255, 0), 3)

        barcodeData = obj.data.decode("utf-8")
        v.add(barcodeData)

        barcodeType = obj.type
        string = "Data " + str(barcodeData) + " | Type " + str(barcodeType)

        cv2.putText(
            frame, string, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2
        )
        # print("Barcode: "+barcodeData +" | Type: "+barcodeType)
        print(barcodeData)


cap = cv2.VideoCapture(0)

v = set()
while True:
    ret, frame = cap.read()
    decoder(frame)
    cv2.imshow("Image", frame)
    code = cv2.waitKey(10)
    if code == ord("q"):
        print(v)
        break
    

wb = load_workbook("scouting.xlsx")

ws = wb.active

for val in v:
    ws.insert_rows(2)
    ws["A2"] = val
    inpl = val.split(";")
    for x in range(2,len(inpl)+2):
        ws.cell(row=2, column=x, value=inpl[x-2]) 


wb.save("scouting.xlsx")
