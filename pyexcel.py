from openpyxl import load_workbook

wb = load_workbook("scouting.xlsx")

ws = wb.active

inp = input("str: ")

ws.insert_rows(2)
ws["A2"] = inp
inpl = inp.split(";")
for x in range(2,len(inpl)+2):
    ws.cell(row=2, column=x, value=inpl[x-2]) 


wb.save("scouting.xlsx")
